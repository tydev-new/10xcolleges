#!/usr/bin/env python3
"""Build the application tracker spreadsheet from a student's meta.json.

Four sheets: Applications, Tasks, Recommenders, Key Dates. The Tasks sheet is generated
by working *backwards* from each college deadline — the whole point of a tracker is that
it tells you what's late in October, not what's due in January.

Calendar facts (when FAFSA opens, the backward plan's offsets) live in
config/calendar.json so they can be corrected without touching code. The *rules* for
computing dates live here, deliberately — see aid_year_anchor().

Usage:
    make_tracker.py students/maya-rodriguez
    make_tracker.py students/maya-rodriguez --today 2026-09-01   # deterministic testing
"""

import argparse
import json
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
except ImportError:
    sys.exit("openpyxl not installed. Run: .venv/bin/pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
CONFIG_PATH = ROOT / "config" / "calendar.json"

HEADER_BG = "1F3864"
HEADER_FG = "FFFFFF"
BAND = "F2F5FA"

RED = PatternFill("solid", start_color="FFC7CE")
AMBER = PatternFill("solid", start_color="FFEB9C")
GREEN = PatternFill("solid", start_color="C6EFCE")
GREY = PatternFill("solid", start_color="E7E6E6")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUSES = ["considering", "researching", "committed-to-apply",
            "in-progress", "submitted", "decided", "withdrawn"]
TASK_STATUSES = ["not started", "in progress", "blocked", "done"]

# Statuses whose applications are finished or abandoned — no task plan needed.
CLOSED_STATUSES = ("withdrawn", "decided", "submitted")


def load_config(path=CONFIG_PATH):
    if not path.exists():
        sys.exit(f"Missing {path}. It holds the calendar facts (FAFSA date, backward plan).")
    return json.loads(path.read_text())


def parse_date(s):
    """Strict ISO parse. Returns None for blank, raises ValueError for malformed.

    Blank and malformed are different problems: a college without a deadline yet is
    normal mid-intake, while '11/01/2026' is a typo that must not silently remove a
    school from the task plan.
    """
    if s is None or str(s).strip() == "":
        return None
    return datetime.strptime(str(s).strip(), "%Y-%m-%d").date()


def aid_year_anchor(deadline, cutoff_month):
    """Return the calendar year in which the FAFSA cycle for this deadline opens.

    Financial aid is keyed to the fall the student matriculates, NOT to the deadline's
    own calendar year. A Regular Decision deadline of 2027-01-15 is for fall 2027 entry,
    whose FAFSA opened in October *2026*. Anchoring on the deadline's year instead
    schedules aid nine months after the applications are due.
    """
    return deadline.year - 1 if deadline.month < cutoff_month else deadline.year


def style_header(ws, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color=HEADER_FG, size=11)
        c.fill = PatternFill("solid", start_color=HEADER_BG)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def finish(ws, n_rows, n_cols):
    """Borders + zebra banding over the data range."""
    for r in range(2, n_rows + 2):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if r % 2 == 0:
                cell.fill = PatternFill("solid", start_color=BAND)


def add_validation(ws, col_letter, options, n_rows):
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{n_rows + 1}")


def read_colleges(meta):
    """Parse deadlines up front so a typo fails loudly instead of dropping a school.

    A malformed deadline used to skip the college's entire task plan silently: the
    school still appeared on the Applications sheet, but none of its twelve tasks were
    generated, so the student had no plan for an application they thought was tracked.
    """
    out, bad = [], []
    for c in meta.get("colleges", []):
        try:
            dl = parse_date(c.get("deadline"))
        except ValueError:
            bad.append((c.get("name", "<unnamed>"), c.get("deadline")))
            continue
        out.append((c, dl))

    if bad:
        lines = "\n".join(f"    {n}: {v!r}" for n, v in bad)
        sys.exit(
            f"Unparseable deadline(s) in meta.json — dates must be ISO (YYYY-MM-DD):\n"
            f"{lines}\n\n"
            "Refusing to build a tracker that would silently omit these schools.\n"
            "Fix meta.json and re-run."
        )
    return out


def sheet_applications(wb, colleges, today):
    ws = wb.create_sheet("Applications")
    headers = ["College", "Tier", "Plan", "Deadline", "Days Left", "App Type",
               "Status", "Essays Req'd", "Essays Done", "Counselor Ltr",
               "Teacher Recs", "Test Policy", "Fee", "Notes"]
    widths = [30, 10, 8, 12, 10, 14, 18, 12, 12, 13, 13, 14, 10, 40]
    style_header(ws, headers, widths)

    order = {"reach": 0, "target": 1, "safety": 2}
    rows = sorted(colleges, key=lambda cd: (cd[1] or date(2099, 1, 1),
                                            order.get(cd[0].get("tier"), 9)))

    for i, (c, dl) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=c.get("name"))
        ws.cell(row=i, column=2, value=c.get("tier"))
        ws.cell(row=i, column=3, value=c.get("decision_plan"))
        d = ws.cell(row=i, column=4, value=dl)
        d.number_format = "yyyy-mm-dd"
        # Live formula, not a baked number: the tracker is meant to be lived in for
        # months, and a static "Days Left" silently lies the day after generation.
        ws.cell(row=i, column=5,
                value=f"=IF(D{i}=\"\",\"\",D{i}-TODAY())" if dl else None)
        ws.cell(row=i, column=6, value=c.get("app_type"))
        ws.cell(row=i, column=7, value=c.get("status", "considering"))
        ws.cell(row=i, column=8, value=c.get("essays_required"))
        ws.cell(row=i, column=9, value=c.get("essays_done", 0))
        ws.cell(row=i, column=10, value="yes" if c.get("counselor_letter") else "no")
        ws.cell(row=i, column=11, value=c.get("teacher_recs"))
        ws.cell(row=i, column=12, value=c.get("test_policy"))
        ws.cell(row=i, column=13, value=c.get("fee"))
        ws.cell(row=i, column=14, value=c.get("notes"))

    n = len(rows)
    finish(ws, n, len(headers))
    if not n:
        return ws

    rng = f"E2:E{n + 1}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan",
                                                  formula=["0"], fill=GREY))
    ws.conditional_formatting.add(rng, CellIsRule(operator="between",
                                                  formula=["0", "13"], fill=RED))
    ws.conditional_formatting.add(rng, CellIsRule(operator="between",
                                                  formula=["14", "29"], fill=AMBER))
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan",
                                                  formula=["29"], fill=GREEN))
    ws.conditional_formatting.add(
        f"A2:N{n + 1}",
        FormulaRule(formula=['$G2="submitted"'], fill=GREEN, stopIfTrue=False),
    )
    add_validation(ws, "G", STATUSES, n)
    add_validation(ws, "B", ["safety", "target", "reach"], n)
    return ws


def build_tasks(colleges, today, cfg):
    """Generate the task list. Returns (tasks, notices) — notices are for the operator."""
    plan = cfg["backward_plan"]
    rules = cfg["rules"]
    span = rules["plan_runway_weeks"]
    tasks, notices = [], []

    for c, dl in colleges:
        name = c.get("name")
        if c.get("status") in CLOSED_STATUSES:
            continue
        if dl is None:
            notices.append(f"{name}: no deadline set — no task plan generated yet.")
            continue

        runway_days = (dl - today).days

        # A deadline that has already passed gets no plan. Generating "SUBMIT
        # application" due today for an application that closed six weeks ago buries
        # the tasks that are still live under a dozen impossible ones.
        if runway_days < 0:
            tasks.append({
                "due": today,
                "task": f"DEADLINE PASSED for {name} ({dl.isoformat()}) — "
                        "confirm submitted, or mark withdrawn in meta.json",
                "college": name, "cat": "Decision", "owner": "Student",
                "notes": f"Deadline was {abs(runway_days)} days ago.",
            })
            notices.append(f"{name}: deadline {dl.isoformat()} already passed.")
            continue

        compressed = 0 <= runway_days < span * 7
        for step in plan:
            weeks = step["weeks_before"]
            if compressed:
                due = dl - timedelta(days=round(weeks / span * runway_days))
            else:
                due = dl - timedelta(weeks=weeks)
            tasks.append({
                "due": due,
                "task": step["task"].format(college=name),
                "college": name,
                "cat": step["category"],
                "owner": step["owner"],
                "notes": "Tight timeline — compressed schedule" if compressed else None,
            })
        if compressed:
            notices.append(f"{name}: {runway_days} days out — plan compressed.")

    dated = [dl for _, dl in colleges if dl]
    if dated:
        earliest = min(dated)
        anchor = aid_year_anchor(earliest, rules["spring_cutoff_month"])
        aid = cfg["aid"]
        fm, fd = (int(x) for x in aid["fafsa_opens"].split("-"))
        cm, cd = (int(x) for x in aid["css_suggested_by"].split("-"))

        tasks.append({"due": date(anchor, fm, fd), "task": "Submit FAFSA (opens Oct 1)",
                      "college": "— all —", "cat": "Aid", "owner": "Family",
                      "notes": aid.get("fafsa_note")})
        tasks.append({"due": date(anchor, cm, cd),
                      "task": "Submit CSS Profile (only if a college requires it)",
                      "college": "— all —", "cat": "Aid", "owner": "Family",
                      "notes": aid.get("css_note")})
        for step in cfg["profile_tasks"]:
            tasks.append({
                "due": earliest - timedelta(weeks=step["weeks_before_earliest"]),
                "task": step["task"], "college": "— all —",
                "cat": step["category"], "owner": step["owner"], "notes": None,
            })

    # A task whose ideal date has passed is still worth doing — surface it as due now
    # rather than as weeks overdue, which just produces a discouraging wall of red.
    catch_up = 0
    for t in tasks:
        if t["due"] < today:
            t["due"] = today
            catch_up += 1
            t["notes"] = t.get("notes") or "Ideal date already passed — do this first"

    tasks.sort(key=lambda t: (t["due"], t["college"]))
    return tasks, notices, catch_up


def sheet_tasks(wb, tasks, today):
    ws = wb.create_sheet("Tasks")
    headers = ["Due", "Days Left", "Task", "College", "Category", "Owner",
               "Status", "Notes"]
    style_header(ws, headers, [12, 10, 52, 26, 12, 10, 14, 34])

    for i, t in enumerate(tasks, start=2):
        d = ws.cell(row=i, column=1, value=t["due"])
        d.number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=2, value=f"=A{i}-TODAY()")
        ws.cell(row=i, column=3, value=t["task"])
        ws.cell(row=i, column=4, value=t["college"])
        ws.cell(row=i, column=5, value=t["cat"])
        ws.cell(row=i, column=6, value=t["owner"])
        ws.cell(row=i, column=7, value="not started")
        ws.cell(row=i, column=8, value=t.get("notes"))

    n = len(tasks)
    finish(ws, n, len(headers))
    if not n:
        return ws

    ws.conditional_formatting.add(
        f"A2:H{n + 1}",
        FormulaRule(formula=['AND($B2<0,$G2<>"done")'], fill=RED, stopIfTrue=True))
    ws.conditional_formatting.add(
        f"A2:H{n + 1}",
        FormulaRule(formula=['$G2="done"'], fill=GREEN, stopIfTrue=True))
    ws.conditional_formatting.add(
        f"A2:H{n + 1}",
        FormulaRule(formula=['AND($B2>=0,$B2<=7)'], fill=AMBER, stopIfTrue=True))
    add_validation(ws, "G", TASK_STATUSES, n)
    ws.auto_filter.ref = f"A1:H{n + 1}"
    return ws


def sheet_recommenders(wb, recs, cfg):
    ws = wb.create_sheet("Recommenders")
    headers = ["Teacher / Counselor", "Subject", "Grade Taught", "Asked On",
               "Agreed?", "Brag Sheet Sent", "Colleges Covered", "Submitted",
               "Thank-You Sent", "Notes"]
    style_header(ws, headers, [24, 16, 13, 12, 10, 15, 30, 12, 14, 28])

    for i, r in enumerate(recs, start=2):
        for col, key in enumerate(
            ["name", "subject", "grade_taught", "asked_on", "agreed",
             "brag_sheet_sent", "colleges", "submitted", "thank_you"], start=1
        ):
            ws.cell(row=i, column=col, value=r.get(key))

    n = max(len(recs), 5)
    finish(ws, n, len(headers))
    for col in ("E", "F", "H", "I"):
        add_validation(ws, col, ["yes", "no"], n)
    return ws


def sheet_key_dates(wb, meta, cfg):
    ws = wb.create_sheet("Key Dates")
    style_header(ws, ["Date", "What", "Notes"], [14, 44, 50])
    rows = meta.get("key_dates") or cfg["key_dates_defaults"]
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r.get("date"))
        ws.cell(row=i, column=2, value=r.get("what"))
        ws.cell(row=i, column=3, value=r.get("notes"))
    finish(ws, len(rows), 3)
    return ws


def sheet_readme(wb, meta, today, cfg):
    ws = wb.create_sheet("Read Me", 0)
    ws.column_dimensions["A"].width = 100
    lead = int(cfg["rules"]["recommender_lead_weeks"])
    lines = [
        (f"{meta.get('name', 'Student')} — Application Tracker", True, 16),
        (f"Generated {today.isoformat()}  •  regenerate any time with the app-tracker skill",
         False, 10),
        ("", False, 11),
        ("How to use this", True, 12),
        ("• Applications — one row per college. Days Left recalculates every time you open "
         "the file: green = comfortable, amber = under 30 days, red = under 14, grey = passed.",
         False, 11),
        ("• Tasks — the real to-do list, built backwards from each deadline. Sort by Due. "
         "Anything red is late. Set Status to 'done' and it turns green.", False, 11),
        (f"• Recommenders — who you asked and whether they've submitted. Ask at least "
         f"{lead} weeks out.", False, 11),
        ("• Key Dates — the fixed calendar: testing, financial aid, reply date.", False, 11),
        ("", False, 11),
        ("Two things worth remembering", True, 12),
        ("1. Deadlines here were copied from each college's own site on the date noted "
         "in your research files. Colleges do change them. Re-check in October.", False, 11),
        ("2. A list without two safeties you'd genuinely be happy attending isn't a "
         "list, it's a wish. Check the Applications tab.", False, 11),
    ]
    for i, (text, bold, size) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=size)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    return ws


def build(meta, today, cfg):
    colleges = read_colleges(meta)
    tasks, notices, catch_up = build_tasks(colleges, today, cfg)

    wb = Workbook()
    wb.remove(wb.active)
    sheet_applications(wb, colleges, today)
    sheet_tasks(wb, tasks, today)
    sheet_recommenders(wb, meta.get("recommenders", []), cfg)
    sheet_key_dates(wb, meta, cfg)
    sheet_readme(wb, meta, today, cfg)
    wb.active = 0
    return wb, tasks, notices, catch_up


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("student_dir")
    p.add_argument("--today", help="ISO date, for deterministic testing")
    p.add_argument("-o", "--out", help="output path (default <student_dir>/out/tracker.xlsx)")
    args = p.parse_args()

    sd = Path(args.student_dir)
    meta_path = sd / "meta.json"
    if not meta_path.exists():
        sys.exit(f"No meta.json in {sd}. The orchestrator should create it first.")
    meta = json.loads(meta_path.read_text())

    try:
        today = parse_date(args.today) or date.today()
    except ValueError:
        sys.exit(f"--today must be ISO (YYYY-MM-DD), got {args.today!r}")

    cfg = load_config()
    wb, tasks, notices, catch_up = build(meta, today, cfg)

    out = Path(args.out) if args.out else sd / "out" / "tracker.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print(f"Wrote {out}")
    print(f"  {len(meta.get('colleges', []))} colleges, {len(tasks)} tasks"
          + (f", {catch_up} needing catch-up as of {today}" if catch_up else ""))
    for n in notices:
        print(f"  note: {n}")


if __name__ == "__main__":
    main()
